
"""
CageMatch Wrestling Scraper  —  Playwright edition
====================================================
Fix summary vs previous version
--------------------------------
* All Playwright calls now run on ONE dedicated worker thread (_worker_thread).
  A Queue(_jobs) serialises work; results come back via tkinter .after().
  This eliminates the "cannot switch to a different thread" crash.
* Matchguide column indices corrected: table has 6 cols
  (#, Date, Fixture, WON, Rating, Votes) → date=tds[1], fixture=tds[2],
  won=tds[3], rating=tds[4].
* WON star text is parsed from raw HTML (latin-1) because BeautifulSoup
  strips the ★ / * characters from <span class="starRating">.
* After promotions load a messagebox asks the user to confirm before
  the browser navigates to the matchguide URLs.
* Navigation goes directly to the constructed URL
  (?id=8&nr=NR&page=7&sortby=colRating&sorttype=DESC) — no UI clicking.

Install:
    pip install playwright playwright-stealth beautifulsoup4 openpyxl
    playwright install chromium

Run GUI:   python scraper.py
Run CLI:   python scraper.py --cli --output out.xlsx --min-rating 8.0
"""

import time, re, sys, argparse, threading, random, queue
from datetime import datetime
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
_DEFAULT_OUTPUT = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                               "wrestling_matches.xlsx")

BASE_URL        = "https://www.cagematch.net"
PROMOTIONS_URL  = f"{BASE_URL}/?id=8&view=promotions&sortby=colRating&sorttype=DESC"
MIN_RATING      = 8.00
REQUEST_DELAY   = 3.5
MAX_RETRIES     = 3

# ──────────────────────────────────────────────────────────────────────────────
# SINGLE-THREAD PLAYWRIGHT WORKER
# All browser operations are serialised through this thread so Playwright's
# sync_api never sees a cross-thread call.
# ──────────────────────────────────────────────────────────────────────────────

_jobs: queue.Queue = queue.Queue()   # (fn, result_future)  posted from any thread
_worker_started    = False
_browser_ready     = threading.Event()

_pw_instance = None
_browser     = None
_context     = None
_page        = None


def _browser_worker():
    """Runs forever on its own thread, executing browser jobs from _jobs."""
    global _pw_instance, _browser, _context, _page
    while True:
        item = _jobs.get()
        if item is None:          # sentinel → shut down
            break
        fn, fut = item
        try:
            result = fn()
            fut["result"] = result
        except Exception as exc:
            fut["error"] = exc
        finally:
            fut["done"].set()


def _submit(fn):
    """Submit *fn* to the browser worker and block until it returns."""
    fut = {"done": threading.Event(), "result": None, "error": None}
    _jobs.put((fn, fut))
    fut["done"].wait()
    if fut["error"] is not None:
        raise fut["error"]
    return fut["result"]


def _ensure_worker():
    global _worker_started
    if not _worker_started:
        t = threading.Thread(target=_browser_worker, daemon=True)
        t.start()
        _worker_started = True


def _start_browser(headless: bool = False):
    def _do():
        global _pw_instance, _browser, _context, _page
        from playwright.sync_api import sync_playwright
        from playwright_stealth import Stealth
        _pw_instance = sync_playwright().start()
        _browser = _pw_instance.chromium.launch(
            headless=headless,
            args=["--disable-blink-features=AutomationControlled",
                  "--no-sandbox", "--disable-setuid-sandbox"],
        )
        _context = _browser.new_context(
            viewport={"width": 1366, "height": 768},
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            locale="en-US",
            timezone_id="America/New_York",
            java_script_enabled=True,
            accept_downloads=False,
            extra_http_headers={"Accept-Language": "en-US,en;q=0.9", "DNT": "1"},
        )
        _page = _context.new_page()
        Stealth().use_sync(_page)
        _page.goto(BASE_URL, wait_until="domcontentloaded", timeout=30_000)
        _human_pause(2, 4)
        _scroll_random()
    _ensure_worker()
    _submit(_do)


def _stop_browser():
    def _do():
        global _pw_instance, _browser, _context, _page
        try:
            if _browser:  _browser.close()
            if _pw_instance: _pw_instance.stop()
        except Exception:
            pass
        _pw_instance = _browser = _context = _page = None
    try:
        _submit(_do)
    except Exception:
        pass
    _jobs.put(None)   # signal worker to exit


# ──────────────────────────────────────────────────────────────────────────────
# LOW-LEVEL HELPERS  (must be called *inside* a _submit lambda)
# ──────────────────────────────────────────────────────────────────────────────

def _human_pause(lo=1.5, hi=3.5):
    time.sleep(random.uniform(lo, hi))

def _scroll_random():
    try:
        dist = random.randint(300, 700)
        _page.mouse.wheel(0, dist)
        time.sleep(random.uniform(0.4, 0.9))
        _page.mouse.wheel(0, -random.randint(50, 150))
    except Exception:
        pass

def _wait_for_captcha_if_needed(log_fn=None):
    title   = (_page.title() or "").lower()
    content = _page.content().lower()
    signals = ["captcha", "cf-challenge", "just a moment", "checking your browser",
               "enable javascript", "ddos-guard", "website security", "click to proceed"]
    if not any(s in title or s in content for s in signals):
        return
    proceed_texts = ["click to proceed", "proceed to page", "i am human", "continue", "proceed"]
    for text in proceed_texts:
        try:
            btn = _page.locator(
                f"button:has-text('{text}'), input[value*='{text}'], a:has-text('{text}')").first
            if btn.is_visible(timeout=2000):
                if log_fn: log_fn(f"  🖱 Auto-clicking: '{text}'")
                btn.click()
                _human_pause(2, 4)
                if not any(s in _page.content().lower() for s in signals):
                    return
        except Exception:
            continue
    msg = "⏸  Security page — solve it in the browser window (auto-continues)…"
    if log_fn: log_fn(msg)
    else: print(msg)
    try:
        _page.wait_for_function(
            "() => !document.body.innerText.toLowerCase().includes('click to proceed')"
            " && !document.body.innerText.toLowerCase().includes('website security')",
            timeout=60_000)
        if log_fn: log_fn("  ✔ Past security page.")
    except Exception:
        if log_fn: log_fn("  ⚠ Timed out — continuing.")


def _get_html(url: str, log_fn=None) -> str:
    """
    Navigate to *url* inside the browser worker thread and return raw HTML.
    Must be called from within a _submit() lambda so it runs on the worker thread.
    """
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            _page.goto(url, wait_until="domcontentloaded", timeout=30_000)
            _wait_for_captcha_if_needed(log_fn=log_fn)
            _human_pause(REQUEST_DELAY, REQUEST_DELAY + 1.5)
            _scroll_random()
            return _page.content()
        except Exception as e:
            wait = 8 * attempt + random.uniform(1, 4)
            if log_fn:
                log_fn(f"  ⚠ Error on attempt {attempt}/{MAX_RETRIES}: {e} — retrying in {wait:.0f}s")
            time.sleep(wait)
    raise RuntimeError(f"Failed to fetch {url} after {MAX_RETRIES} attempts.")


def get_soup(url: str, log_fn=None) -> BeautifulSoup:
    html = _submit(lambda: _get_html(url, log_fn=log_fn))
    return BeautifulSoup(html, "html.parser")


# ──────────────────────────────────────────────────────────────────────────────
# PROMOTIONS SCRAPER
# ──────────────────────────────────────────────────────────────────────────────

def scrape_promotions_page(url: str, log_fn=None):
    soup = get_soup(url, log_fn=log_fn)
    table = soup.find("table", class_="TBase")
    if not table:
        return [], None

    promotions = []
    for row in table.find_all("tr")[1:]:
        tds = row.find_all("td")
        if len(tds) < 6:
            continue
        name_link = tds[2].find("a")
        if not name_link:
            continue
        name     = name_link.get_text(strip=True)
        href     = name_link.get("href", "")
        location = tds[3].get_text(strip=True)
        years    = tds[4].get_text(strip=True)
        rating_span = tds[5].find("span", class_="Rating")
        if not rating_span:
            continue
        try:
            rating = float(rating_span.get_text(strip=True))
        except ValueError:
            continue
        nr_m = re.search(r"nr=(\d+)", href)
        nr   = nr_m.group(1) if nr_m else None
        profile_url = f"{BASE_URL}/{href}" if href.startswith("?") else href
        promotions.append({
            "name": name, "location": location, "years": years,
            "rating": rating, "profile_url": profile_url, "nr": nr,
        })

    next_url = None
    for a in soup.find_all("a"):
        div = a.find("div")
        if div and div.get_text(strip=True) == ">":
            h = a.get("href", "")
            next_url = f"{BASE_URL}/{h}" if h.startswith("?") else h
            break

    return promotions, next_url


def fetch_all_promotions(min_rating: float = MIN_RATING, log_fn=None) -> list:
    url, page_num, all_promos = PROMOTIONS_URL, 1, []
    while url:
        if log_fn: log_fn(f"  Fetching promotions page {page_num} …")
        promotions, next_url = scrape_promotions_page(url, log_fn=log_fn)
        if not promotions:
            break
        page_min  = min(p["rating"] for p in promotions)
        qualified = [p for p in promotions if p["rating"] >= min_rating]
        all_promos.extend(qualified)
        if log_fn: log_fn(f"    {len(qualified)} qualifying (page min: {page_min:.2f})")
        if page_min < min_rating:
            break
        url = next_url
        page_num += 1
    return sorted(all_promos, key=lambda x: x["rating"], reverse=True)


# ──────────────────────────────────────────────────────────────────────────────
# MATCHGUIDE SCRAPER
# Table columns: # | Date | Match fixture | WON | Rating | Votes
#                0     1         2           3      4       5
# ──────────────────────────────────────────────────────────────────────────────
def get_matchguide_url(nr: str) -> str:
    return f"{BASE_URL}/?id=8&nr={nr}&page=7&sortby=colRating&sorttype=DESC"

def get_matchguide_won_url(nr: str) -> str:
    return f"{BASE_URL}/?id=8&nr={nr}&page=7&sortby=colMeltzer&sorttype=DESC"

def _parse_won_from_html(td_html: str) -> str:
    """
    Extract WON star rating text from raw td HTML.
    BeautifulSoup strips the ★/*  characters; we read the raw bytes instead.
    Handles patterns like: ****1/4  ***3/4  **  ½  etc.
    """
    # Try BS first (works when text is preserved)
    from bs4 import BeautifulSoup as _BS
    td = _BS(td_html, "html.parser").find("span", class_="starRating")
    if td:
        t = td.get("title", "").strip() or td.get_text(strip=True)
        if t:
            return t
    # Fallback: grab text between the span tags from raw HTML
    m = re.search(r'class="starRating"[^>]*>(.*?)</span>', td_html, re.DOTALL)
    if m:
        raw = m.group(1).strip()
        if raw:
            return raw
    return ""


def scrape_matchguide_page(url: str, log_fn=None):
    """Fetch one matchguide page; returns (matches, next_url)."""
    # We need both the parsed soup AND raw HTML for WON extraction
    def _fetch_both():
        html = _get_html(url, log_fn=log_fn)
        return html, BeautifulSoup(html, "html.parser")

    html, soup = _submit(_fetch_both)

    table = soup.find("table", class_="TBase")
    if not table:
        return [], None

    # Build a parallel list of raw <td> HTML strings for WON column
    raw_rows = re.findall(r'<tr[^>]*class="TRow\d"[^>]*>(.*?)</tr>', html, re.DOTALL | re.IGNORECASE)

    matches = []
    soup_rows = [r for r in table.find_all("tr") if r.get("class") and
                 any(c.startswith("TRow") for c in r.get("class", []))]

    for idx, row in enumerate(soup_rows):
        tds = row.find_all("td")
        if len(tds) < 5:
            continue
        # Columns: 0=#  1=Date  2=Fixture  3=WON  4=Rating  5=Votes
        date         = tds[1].get_text(strip=True)
        fixture_link = tds[2].find("a")
        fixture      = fixture_link.get_text(strip=True) if fixture_link else tds[2].get_text(strip=True)
        match_url    = ""
        if fixture_link:
            h = fixture_link.get("href", "")
            match_url = f"{BASE_URL}/{h}" if h.startswith("?") else h

        # WON: try soup first, fall back to raw HTML
        won = tds[3].find("span", class_="starRating")
        won_text = ""
        if won:
            won_text = (won.get("title", "").strip()
                        or won.get_text(strip=True))
        if not won_text and idx < len(raw_rows):
            won_text = _parse_won_from_html(raw_rows[idx])

        rating_span = tds[4].find("span", class_="Rating")
        rating_str, rating_val = "", None
        if rating_span:
            rating_str = rating_span.get_text(strip=True)
            try:
                rating_val = float(rating_str)
            except ValueError:
                pass

        matches.append({
            "date": date, "fixture": fixture, "match_url": match_url,
            "won": won_text, "rating": rating_str, "rating_val": rating_val,
        })

    next_url = None
    for a in soup.find_all("a"):
        div = a.find("div")
        if div and div.get_text(strip=True) == ">":
            h = a.get("href", "")
            next_url = f"{BASE_URL}/{h}" if h.startswith("?") else h
            break

    return matches, next_url


def fetch_matches_for_promotion(nr: str, min_rating: float = None,
                                 min_won: float = None, log_fn=None) -> tuple:
    """
    Returns (rating_matches, won_only_matches).

    rating_matches : rows that meet min_rating (first priority).
    won_only_matches: rows that meet min_won but are NOT already in rating_matches,
                      collected from the WON-sorted pass.

    Either filter may be None (skip that pass).
    """
    rating_matches = []
    rating_match_keys = set()   # (date, fixture) dedup keys

    # ── Pass 1: sort by Cagematch Rating ────────────────────────────────────
    if min_rating is not None:
        url, page_num = get_matchguide_url(nr), 1
        while url:
            if log_fn: log_fn(f"    [Rating pass] Page {page_num}: {url}")
            matches, next_url = scrape_matchguide_page(url, log_fn=log_fn)
            if not matches:
                break
            qualified = [m for m in matches
                         if m["rating_val"] is not None and m["rating_val"] >= min_rating]
            rating_matches.extend(qualified)
            for m in qualified:
                rating_match_keys.add((m["date"], m["fixture"]))
            if log_fn: log_fn(f"      {len(qualified)} qualifying")
            rated = [m for m in matches if m["rating_val"] is not None]
            if rated and max(m["rating_val"] for m in rated) < min_rating:
                break
            if not rated and not next_url:
                break
            url = next_url
            page_num += 1

    # ── Pass 2: sort by WON stars ────────────────────────────────────────────
    won_only_matches = []
    if min_won is not None:
        url, page_num = get_matchguide_won_url(nr), 1
        while url:
            if log_fn: log_fn(f"    [WON pass]    Page {page_num}: {url}")
            matches, next_url = scrape_matchguide_page(url, log_fn=log_fn)
            if not matches:
                break
            for m in matches:
                wv = _won_to_float(m["won"])
                if wv is None or wv < min_won:
                    # Since sorted DESC, once we drop below threshold stop
                    next_url = None
                    break
                key = (m["date"], m["fixture"])
                if key not in rating_match_keys:   # not already in table 1
                    won_only_matches.append(m)
            if log_fn: log_fn(f"      {len(won_only_matches)} WON-only so far")
            url = next_url
            page_num += 1

    return rating_matches, won_only_matches

# ──────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ──────────────────────────────────────────────────────────────────────────────

_HDR_FILL  = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
_ALT_FILL  = PatternFill("solid", start_color="D6E4F0", end_color="D6E4F0")
_HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
_DATA_FONT = Font(name="Arial", size=10)
_BORDER    = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
_RATING_COLORS = {(9.0, 99): "00B050", (8.5, 9.0): "92D050", (8.0, 8.5): "FFFF00"}


def _rcolor(val):
    if val is None: return None
    for (lo, hi), c in _RATING_COLORS.items():
        if lo <= val < hi: return c
    if val >= 9.0: return "00B050"
    return None


def _safe_sheet(name: str) -> str:
    for ch in r'\/:*?"<>|[]':
        name = name.replace(ch, "")
    return name[:31]


def _parse_dt(s):
    try:
        return datetime.strptime(s.strip(), "%d.%m.%Y")
    except Exception:
        return None


def _won_to_float(s):
    """Convert WON star string to float.
    Accepts: '****1/2', '***3/4', '**', '4.5', '5', '★★★½', etc.
    """
    if not s:
        return None
    s = s.strip()
    # Plain numeric first
    try:
        return float(s)
    except ValueError:
        pass
    # Count * or ★ characters
    cnt = s.count("*") + s.count("\u2605")
    if cnt == 0:
        return None
    # Fraction suffixes
    frac = 0.0
    if "3/4" in s or "\u00be" in s:
        frac = 0.75
    elif "1/2" in s or "\u00bd" in s:
        frac = 0.5
    elif "1/4" in s or "\u00bc" in s:
        frac = 0.25
    return float(cnt) + frac

def _col_letter(n: int) -> str:
    """Convert 1-based column index to Excel letter (1→A, 6→F, etc.)."""
    result = ""
    while n:
        n, r = divmod(n - 1, 26)
        result = chr(65 + r) + result
    return result

def export_to_excel(data: dict, output_path: str,
                    date_from=None, date_to=None,
                    won_filter=None, rating_filter=None):
    """
    data: dict of promotion_name -> (rating_matches, won_only_matches)

    Layout per sheet:
      Table 1 starts at col A  — header "Cagematch Rating" merged across A:D
      Table 2 starts at col F  — header "Only Has WON Stars" merged across F:I
      (one blank column E between them)

    Rows flow independently; each table has its own Date/Fixture/WON/Rating sub-headers.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    dt_from = _parse_dt(date_from) if date_from else None
    dt_to   = _parse_dt(date_to)   if date_to   else None
    won_min = _won_to_float(won_filter) if won_filter else None

    # Column offsets: Table 1 = cols 1-4 (A-D), Table 2 = cols 6-9 (F-I)
    T1_START = 1
    T2_START = 6

    SUB_HDRS = ["Date", "Match Fixture", "WON Stars", "Rating"]

    _MERGED_FILL_1 = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    _MERGED_FILL_2 = PatternFill("solid", start_color="7B2D8B", end_color="7B2D8B")
    _SUBHDR_FILL   = PatternFill("solid", start_color="2E75B6", end_color="2E75B6")
    _SUBHDR_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    _MERGED_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=12)

    def _apply_date_won_filters(matches):
        out = []
        for m in matches:
            dt = _parse_dt(m["date"])
            if dt_from and dt and dt < dt_from: continue
            if dt_to   and dt and dt > dt_to:   continue
            out.append(m)
        return out

    def _write_merged_header(ws, col_start, text, fill):
        col_end = col_start + 3
        ws.merge_cells(start_row=1, start_column=col_start,
                       end_row=1,   end_column=col_end)
        c = ws.cell(row=1, column=col_start, value=text)
        c.font      = _MERGED_FONT
        c.fill      = fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _BORDER

    def _write_sub_headers(ws, col_start):
        for i, h in enumerate(SUB_HDRS):
            c = ws.cell(row=2, column=col_start + i, value=h)
            c.font      = _SUBHDR_FONT
            c.fill      = _SUBHDR_FILL
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = _BORDER
        ws.row_dimensions[2].height = 20

    def _write_match_row(ws, row_num, col_start, m, alt):
        def wc(offset, value, link=None, align="left", nfmt=None):
            c = ws.cell(row=row_num, column=col_start + offset, value=value)
            c.font      = _DATA_FONT
            c.border    = _BORDER
            c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
            if alt: c.fill = alt
            if link:
                c.hyperlink = link
                c.font = Font(name="Arial", size=10, color="0563C1", underline="single")
            if nfmt: c.number_format = nfmt
            return c

        wc(0, m["date"], align="center")
        wc(1, m["fixture"], link=m.get("match_url") or None)
        wc(2, m["won"], align="center")
        rv  = m["rating_val"]
        rc  = wc(3, rv if rv is not None else m["rating"], align="center", nfmt="0.00")
        clr = _rcolor(rv)
        if clr: rc.fill = PatternFill("solid", start_color=clr, end_color=clr)

    for promo, (rating_matches, won_only_matches) in data.items():
        ws = wb.create_sheet(title=_safe_sheet(promo))
        ws.row_dimensions[1].height = 24

        # ── Merged section headers (row 1) ───────────────────────────────────
        _write_merged_header(ws, T1_START, "Cagematch Rating",   _MERGED_FILL_1)
        _write_merged_header(ws, T2_START, "Only Has WON Stars", _MERGED_FILL_2)

        # ── Sub-headers (row 2) ──────────────────────────────────────────────
        _write_sub_headers(ws, T1_START)
        _write_sub_headers(ws, T2_START)

        # ── Table 1: rating matches ──────────────────────────────────────────
        t1_rows = _apply_date_won_filters(rating_matches)
        row_num = 3
        for m in t1_rows:
            alt = _ALT_FILL if row_num % 2 == 0 else None
            _write_match_row(ws, row_num, T1_START, m, alt)
            row_num += 1
        t1_last_data_row = row_num - 1

        # Summary row for Table 1
        if t1_last_data_row >= 3:
            sr = t1_last_data_row + 2
            ws.cell(row=sr, column=T1_START,     value="Total Matches:").font = Font(bold=True, name="Arial")
            ws.cell(row=sr, column=T1_START + 1, value=t1_last_data_row - 2).font = Font(name="Arial")
            ws.cell(row=sr, column=T1_START + 2, value="Avg Rating:").font = Font(bold=True, name="Arial")
            ac = ws.cell(row=sr, column=T1_START + 3,
                         value=f'=IFERROR(AVERAGE(D3:D{t1_last_data_row}),"")')
            ac.number_format = "0.00"
            ac.font = Font(name="Arial")

        # ── Table 2: WON-only matches ────────────────────────────────────────
        t2_rows = _apply_date_won_filters(won_only_matches)
        row_num = 3
        for m in t2_rows:
            alt = _ALT_FILL if row_num % 2 == 0 else None
            _write_match_row(ws, row_num, T2_START, m, alt)
            row_num += 1
        t2_last_data_row = row_num - 1

        # Summary row for Table 2
        # Excel col letters for T2: F=6 G=7 H=8 I=9
        col_i_letter = "I"
        if t2_last_data_row >= 3:
            sr2 = t2_last_data_row + 2
            ws.cell(row=sr2, column=T2_START,     value="Total Matches:").font = Font(bold=True, name="Arial")
            ws.cell(row=sr2, column=T2_START + 1, value=t2_last_data_row - 2).font = Font(name="Arial")
            ws.cell(row=sr2, column=T2_START + 2, value="Avg Rating:").font = Font(bold=True, name="Arial")
            ac2 = ws.cell(row=sr2, column=T2_START + 3,
                          value=f'=IFERROR(AVERAGE({col_i_letter}3:{col_i_letter}{t2_last_data_row}),"")')
            ac2.number_format = "0.00"
            ac2.font = Font(name="Arial")

        # ── Column widths ────────────────────────────────────────────────────
        for base in (T1_START, T2_START):
            ws.column_dimensions[_col_letter(base)].width     = 14   # Date
            ws.column_dimensions[_col_letter(base+1)].width   = 52   # Fixture
            ws.column_dimensions[_col_letter(base+2)].width   = 14   # WON
            ws.column_dimensions[_col_letter(base+3)].width   = 12   # Rating
        # Blank spacer col E
        ws.column_dimensions["E"].width = 3

        ws.freeze_panes = "A3"

    wb.save(output_path)
    return output_path

# ──────────────────────────────────────────────────────────────────────────────
# CLI
# ──────────────────────────────────────────────────────────────────────────────

def run_cli(args):
    log = lambda m: print(m, flush=True)
    min_r    = args.min_rating if args.min_rating else None
    won_min  = _won_to_float(args.won_min) if args.won_min else None
    _start_browser(headless=False)
    try:
        log(f"Fetching promotions (min rating={min_r}, min WON={won_min}) …")
        promos = fetch_all_promotions(min_rating=min_r or MIN_RATING, log_fn=log)
        log(f"\nFound {len(promos)} qualifying promotions.\n")
        for i, p in enumerate(promos, 1):
            print(f"  {i:3}. [{p['rating']:.2f}]  {p['name']}  ({p['location']})  {p['years']}")
        data = {}
        for p in promos:
            if not p["nr"]: continue
            log(f"\nScraping: {p['name']}  (nr={p['nr']})")
            r_m, w_m = fetch_matches_for_promotion(
                p["nr"], min_rating=min_r, min_won=won_min, log_fn=log)
            log(f"  → {len(r_m)} rating + {len(w_m)} WON-only matches")
            data[p["name"]] = (r_m, w_m)
        log(f"\nWriting {args.output} …")
        export_to_excel(data, args.output,
                        date_from=args.date_from, date_to=args.date_to,
                        won_filter=args.won_min, rating_filter=min_r)
        log(f"✅ Done — {args.output}")
    finally:
        _stop_browser()


# ──────────────────────────────────────────────────────────────────────────────
# GUI
# ──────────────────────────────────────────────────────────────────────────────

def run_gui():
    import tkinter as tk
    from tkinter import ttk, messagebox
    from tkinter.filedialog import asksaveasfilename

    class App(tk.Tk):
        def __init__(self):
            super().__init__()
            self.title("CageMatch Wrestling Scraper")
            self.geometry("1060x730")
            self.resizable(True, True)
            self.configure(bg="#1a1a2e")
            self.promotions = []
            self._build_ui()
            self.protocol("WM_DELETE_WINDOW", self._on_close)

        def _on_close(self):
            _stop_browser()
            self.destroy()

        def _build_ui(self):
            s = ttk.Style(self)
            s.theme_use("clam")
            s.configure("TFrame",    background="#1a1a2e")
            s.configure("TLabel",    background="#1a1a2e", foreground="#e0e0e0", font=("Arial", 10))
            s.configure("H.TLabel",  background="#1a1a2e", foreground="#c9a227", font=("Arial", 14, "bold"))
            s.configure("TButton",   font=("Arial", 10, "bold"), padding=6)
            s.configure("Treeview",  font=("Arial", 10), rowheight=24,
                        background="#16213e", foreground="#e0e0e0", fieldbackground="#16213e")
            s.configure("Treeview.Heading", font=("Arial", 10, "bold"),
                        background="#0f3460", foreground="#c9a227")
            s.map("Treeview",
                  background=[("selected", "#c9a227")],
                  foreground=[("selected", "#1a1a2e")])

            ttk.Label(self, text="⚡  CageMatch Wrestling Scraper", style="H.TLabel").pack(pady=(14, 2))
            ttk.Label(self, text="cagematch.net — promotions rated ≥ 8.00", foreground="#888").pack()

            nb = ttk.Notebook(self)
            nb.pack(fill="both", expand=True, padx=14, pady=10)

            t1, t2, tL = ttk.Frame(nb), ttk.Frame(nb), ttk.Frame(nb)
            nb.add(t1, text="  Step 1 · Promotions  ")
            nb.add(t2, text="  Step 2 · Filters & Export  ")
            nb.add(tL, text="  Log  ")

            self._tab1(t1)
            self._tab2(t2)
            self._tab_log(tL)

        # ── Tab 1 ─────────────────────────────────────────────────────────────

        def _tab1(self, f):
            ctrl = ttk.Frame(f); ctrl.pack(fill="x", padx=8, pady=8)
            ttk.Label(ctrl, text="Min Rating:").pack(side="left")
            self.min_r = tk.StringVar(value="8.00")
            ttk.Entry(ctrl, textvariable=self.min_r, width=6).pack(side="left", padx=4)
            self.fetch_btn = ttk.Button(ctrl, text="🔄  Fetch Promotions", command=self._fetch)
            self.fetch_btn.pack(side="left", padx=10)
            self.sel_all = tk.BooleanVar(value=True)
            ttk.Checkbutton(ctrl, text="Select All", variable=self.sel_all,
                            command=self._toggle_all).pack(side="left")
            self.cnt_lbl = ttk.Label(ctrl, text="", foreground="#c9a227")
            self.cnt_lbl.pack(side="right", padx=8)

            tbl = ttk.Frame(f); tbl.pack(fill="both", expand=True, padx=8, pady=(0, 8))
            cols = ("sel", "name", "location", "years", "rating")
            self.tree = ttk.Treeview(tbl, columns=cols, show="headings", selectmode="browse")
            for col, hdr, w, anchor in [
                ("sel",      "☑",             38,  "center"),
                ("name",     "Promotion Name", 360, "w"),
                ("location", "Location",       200, "w"),
                ("years",    "Years",          120, "center"),
                ("rating",   "Rating ▼",        80, "center"),
            ]:
                self.tree.heading(col, text=hdr)
                self.tree.column(col, width=w, anchor=anchor, stretch=(col == "name"))
            vsb = ttk.Scrollbar(tbl, orient="vertical", command=self.tree.yview)
            self.tree.configure(yscrollcommand=vsb.set)
            self.tree.pack(side="left", fill="both", expand=True)
            vsb.pack(side="right", fill="y")
            self.tree.tag_configure("on",  background="#1c3a1c", foreground="#90ee90")
            self.tree.tag_configure("off", background="#16213e", foreground="#aaaaaa")
            self.tree.bind("<ButtonRelease-1>", self._click)

        # ── Tab 2 ─────────────────────────────────────────────────────────────

        def _tab2(self, f):
            P = {"padx": 14, "pady": 4}
            def sec(txt):
                ttk.Label(f, text=txt, foreground="#c9a227").pack(anchor="w", padx=14, pady=(12, 0))

            sec("── Date Range Filter (DD.MM.YYYY) ──")
            dr = ttk.Frame(f); dr.pack(fill="x", **P)
            ttk.Label(dr, text="From:").pack(side="left")
            self.dfrom = tk.StringVar()
            ttk.Entry(dr, textvariable=self.dfrom, width=13).pack(side="left", padx=4)
            ttk.Label(dr, text="  To:").pack(side="left")
            self.dto = tk.StringVar()
            ttk.Entry(dr, textvariable=self.dto, width=13).pack(side="left", padx=4)
            ttk.Label(dr, text="blank = no limit", foreground="#777").pack(side="left", padx=8)

            sec("── WON Star Rating Filter ──")
            wr = ttk.Frame(f); wr.pack(fill="x", **P)
            ttk.Label(wr, text="Min WON Stars:").pack(side="left")
            self.won = tk.StringVar()
            ttk.Entry(wr, textvariable=self.won, width=8).pack(side="left", padx=4)
            ttk.Label(wr, text="e.g. 4.5   blank = no filter", foreground="#777").pack(side="left", padx=6)

            sec("── Minimum Match Rating ──")
            mr = ttk.Frame(f); mr.pack(fill="x", **P)
            ttk.Label(mr, text="Min Rating:").pack(side="left")
            self.mrat = tk.StringVar(value="8.00")
            ttk.Entry(mr, textvariable=self.mrat, width=8).pack(side="left", padx=4)

            ttk.Separator(f, orient="horizontal").pack(fill="x", padx=14, pady=14)

            op = ttk.Frame(f); op.pack(fill="x", padx=14)
            ttk.Label(op, text="Output File:").pack(side="left")
            self.out = tk.StringVar(value=_DEFAULT_OUTPUT)
            ttk.Entry(op, textvariable=self.out, width=44).pack(side="left", padx=4)
            ttk.Button(op, text="Browse…", command=self._browse).pack(side="left", padx=4)

            ttk.Separator(f, orient="horizontal").pack(fill="x", padx=14, pady=14)

            self.run_btn = ttk.Button(f, text="⚡  Scrape Selected & Export to Excel",
                                      command=self._run)
            self.run_btn.pack(pady=6)
            self.prog = ttk.Progressbar(f, mode="indeterminate", length=420)
            self.prog.pack(pady=4)
            self.stat = ttk.Label(f, text="", foreground="#c9a227")
            self.stat.pack()

        # ── Log tab ───────────────────────────────────────────────────────────

        def _tab_log(self, f):
            self.log_box = tk.Text(f, bg="#060610", fg="#00ff41",
                                   font=("Courier", 9), state="disabled", wrap="word")
            vsb = ttk.Scrollbar(f, orient="vertical", command=self.log_box.yview)
            self.log_box.configure(yscrollcommand=vsb.set)
            self.log_box.pack(side="left", fill="both", expand=True)
            vsb.pack(side="right", fill="y")

        # ── Helpers ───────────────────────────────────────────────────────────

        def _log(self, msg):
            self.log_box.configure(state="normal")
            self.log_box.insert("end", msg + "\n")
            self.log_box.see("end")
            self.log_box.configure(state="disabled")
            self.update_idletasks()

        def _toggle_all(self):
            sym = "☑" if self.sel_all.get() else "☐"
            tag = "on"  if self.sel_all.get() else "off"
            for item in self.tree.get_children():
                v = list(self.tree.item(item, "values"))
                v[0] = sym
                self.tree.item(item, values=v, tags=(tag,))

        def _click(self, event):
            if (self.tree.identify("region", event.x, event.y) == "cell"
                    and self.tree.identify_column(event.x) == "#1"):
                item = self.tree.identify_row(event.y)
                if item:
                    v = list(self.tree.item(item, "values"))
                    v[0] = "☐" if v[0] == "☑" else "☑"
                    tag  = "on" if v[0] == "☑" else "off"
                    self.tree.item(item, values=v, tags=(tag,))

        # ── Step 1: Fetch promotions ──────────────────────────────────────────

        def _fetch(self):
            self.fetch_btn.configure(state="disabled")
            self._log("=" * 60)
            self._log("Starting browser …")
            try:
                min_r = float(self.min_r.get())
            except ValueError:
                min_r = MIN_RATING

            def run():
                try:
                    _start_browser(headless=False)
                    promos = fetch_all_promotions(min_rating=min_r, log_fn=self._log)
                    self.after(0, lambda: self._populate(promos))
                except Exception as e:
                    msg = str(e)
                    self.after(0, lambda: self._fetch_err(msg))

            threading.Thread(target=run, daemon=True).start()

        def _fetch_err(self, msg):
            self._log(f"❌ Error: {msg}")
            self.fetch_btn.configure(state="normal")
            from tkinter import messagebox
            messagebox.showerror("Fetch failed", msg)

        def _populate(self, promos):
            self.promotions = promos
            self.tree.delete(*self.tree.get_children())
            for p in promos:
                self.tree.insert("", "end",
                                 values=("☑", p["name"], p["location"],
                                         p["years"], f"{p['rating']:.2f}"),
                                 tags=("on",))
            self.cnt_lbl.configure(text=f"{len(promos)} promotions")
            self._log(f"✅ {len(promos)} qualifying promotions loaded.")
            self.fetch_btn.configure(state="normal")

            # ── Prompt user before scraping matchguides ────────────────────
            from tkinter import messagebox
            sel = self._selected()
            if not sel:
                messagebox.showinfo(
                    "Promotions loaded",
                    f"{len(promos)} promotions loaded.\n\n"
                    "Tick the ones you want, set filters on Step 2, then click\n"
                    "⚡ Scrape Selected & Export to Excel."
                )
                return

            ans = messagebox.askyesno(
                "Promotions loaded — proceed to scrape?",
                f"{len(promos)} qualifying promotions found.\n\n"
                f"{len(sel)} are currently selected (all ticked by default).\n\n"
                "The browser will now navigate to each promotion's Matchguide\n"
                "(sorted by rating) and collect matches.\n\n"
                "Proceed?  (You can also untick promotions first, then use the\n"
                "⚡ button on Step 2 to start manually.)",
            )
            if ans:
                self._kick_off_scrape()

        # ── Step 2 / manual kick-off ──────────────────────────────────────────

        def _browse(self):
            p = asksaveasfilename(defaultextension=".xlsx",
                                  filetypes=[("Excel files", "*.xlsx")])
            if p: self.out.set(p)

        def _selected(self):
            sel = []
            for i, item in enumerate(self.tree.get_children()):
                if self.tree.item(item, "values")[0] == "☑" and i < len(self.promotions):
                    sel.append(self.promotions[i])
            return sel

        def _run(self):
            """Manual 'Scrape' button on Step 2."""
            sel = self._selected()
            if not sel:
                from tkinter import messagebox
                messagebox.showwarning("Nothing selected", "Tick at least one promotion.")
                return
            out = self.out.get().strip()
            if not out:
                from tkinter import messagebox
                messagebox.showwarning("No output path", "Set an output file path.")
                return
            self._kick_off_scrape()

        def _kick_off_scrape(self):
            sel = self._selected()
            if not sel:
                from tkinter import messagebox
                messagebox.showwarning("Nothing selected", "Tick at least one promotion.")
                return
            out = self.out.get().strip() or "wrestling_matches.xlsx"
            try:
                mr = float(self.mrat.get())
            except ValueError:
                mr = MIN_RATING

            self.run_btn.configure(state="disabled")
            self.prog.start(10)
            self.stat.configure(text="Scraping …")

            threading.Thread(
                target=self._do_scrape,
                args=(sel, out,
                      self.dfrom.get().strip() or None,
                      self.dto.get().strip()   or None,
                      self.won.get().strip()   or None,
                      mr),
                daemon=True,
            ).start()

        # ── Scrape worker (background thread — only calls _submit, not browser directly) ──

        def _do_scrape(self, sel, out, df, dt, wf, mr):
            self._log("=" * 60)
            won_min = _won_to_float(wf) if wf else None
            self._log(
                f"Scraping {len(sel)} promotion(s)  |  "
                f"min rating: {mr if mr is not None else '—'}  |  "
                f"min WON: {won_min if won_min is not None else '—'}"
            )
            data = {}
            for p in sel:
                self._log(f"\n▶  {p['name']}  (nr={p['nr']})")
                if not p["nr"]:
                    self._log("  ⚠ No nr — skipped")
                    continue
                try:
                    r_matches, w_matches = fetch_matches_for_promotion(
                        p["nr"],
                        min_rating=mr    if mr    else None,
                        min_won   =won_min if won_min else None,
                        log_fn=self._log,
                    )
                    self._log(f"  ✔ {len(r_matches)} rating matches, {len(w_matches)} WON-only matches")
                    data[p["name"]] = (r_matches, w_matches)
                except Exception as e:
                    self._log(f"  ❌ Error: {e}")
                    data[p["name"]] = ([], [])

            self._log("\nWriting Excel workbook …")
            try:
                export_to_excel(data, out, date_from=df, date_to=dt,
                                won_filter=wf, rating_filter=mr)
                self._log(f"✅ Saved: {out}")
                self.after(0, self._done, out)
            except Exception as e:
                self._log(f"❌ Export error: {e}")
                self.after(0, lambda: messagebox.showerror("Export failed", str(e)))
                self.after(0, self._reset)

        def _done(self, out):
            self.prog.stop()
            self.run_btn.configure(state="normal")
            self.stat.configure(text=f"✅  Saved: {out}")
            from tkinter import messagebox
            messagebox.showinfo("Done!", f"Export complete:\n{out}")

        def _reset(self):
            self.prog.stop()
            self.run_btn.configure(state="normal")
            self.stat.configure(text="")

    App().mainloop()


# ──────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ──────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="CageMatch Wrestling Scraper")
    ap.add_argument("--cli",        action="store_true")
    ap.add_argument("--output", default=_DEFAULT_OUTPUT)
    ap.add_argument("--min-rating", type=float, default=8.0)
    ap.add_argument("--date-from",  default=None)
    ap.add_argument("--date-to",    default=None)
    ap.add_argument("--won-min",    default=None)
    args = ap.parse_args()

    if args.cli:
        run_cli(args)
    else:
        run_gui()






